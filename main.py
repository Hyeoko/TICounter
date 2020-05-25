from board import Board
from bet import Bet

# import kivy
from kivy.app import App
# from kivy.uix.label import Label
# from kivy.uix.gridlayout import GridLayout
# from kivy.uix.scrollview import ScrollView
# from kivy.uix.button import Button
from kivy.uix.widget import Widget
from kivy.properties import ObjectProperty
from kivy.uix.popup import Popup

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Color, Font

filename = "Practice.xlsx"

# Shits to add and fix

# Add running totals to print in excel sheet
# Figure out how to record tendencies in excel and in app


class Screen(Widget):
    myBet = ObjectProperty(None) # This is used for Excel
    mySide = ObjectProperty(None)

    # total_profit = ObjectProperty(None)
    micro_game_profit = ObjectProperty(None)
    money_count = ObjectProperty(None)
    current_bet = ObjectProperty(None)

    total_wins = ObjectProperty(None)

    total_bets = ObjectProperty(None)
    shoe_ti = ObjectProperty(None)
    pre_ti = ObjectProperty(None)

    board = ObjectProperty(None)

    def reset(self):
        Bet.reset()
        Board.reset()
        self.myBet = ''
        self.mySide = ''
        # self.total_profit.text = "Total Profit: 0"
        self.micro_game_profit.text = "Micro-game Profit: 0"
        self.money_count = "Money Count: 0"
        self.current_bet.text = "? bet on ?"
        self.total_wins.text = "P: 0     B: 0"
        self.total_bets.text = "# of bets: 0"
        self.shoe_ti.text = "Shoe TI: 0"
        self.pre_ti.text = "Pre-TI: None"
        self.board.text = ''

    def bet_pressed(self, bet, ind=None):
        if bet == 'B':
            Bet.base_bet()
            self.myBet = 'B'
            self.current_bet.text = 'B bet on ?'
        elif bet == "F":
            Bet.force_bet(ind)
            self.myBet = Bet.theForceText[ind]
            self.current_bet.text = Bet.theForceText[ind] + ' bet on ?'
        elif bet == "L":
            Bet.ladder_bet()
            self.myBet = 'L' + str(Bet.get_current_bet())
            self.current_bet.text = 'L' + str(Bet.get_current_bet()) + ' bet on ?'
        else:
            pass

    def end_micro_game(self):
        if len(Bet.profitHistory) != 0:
            Bet.profit_reset()
            self.micro_game_profit.text = 'Micro-game Profit: ' + str(Bet.micro_game_profit)

    def bet_side(self, side):
        if self.myBet is None:
            print("You must pick your bet amount!")
        elif side == 'c':
            self.mySide = ''
            self.current_bet.text = self.myBet + ' bet on ?'
        else:
            self.mySide = Bet.bet_side(side)
            self.current_bet.text = self.myBet + ' bet on ' + side

    def board_pressed(self, winner):
        last_winner = Board.winner(winner)
        checkmark = ''
        amount = ''

        # If bet was placed
        if self.mySide == 'P' or self.mySide == 'B':
            if Bet.bet_result(self.mySide == last_winner):
                checkmark = '+'
            else:
                checkmark = 'x'
            Bet.placeHistory.append(len(Board.history) + 1)
            Bet.sideHistory.append(self.mySide)
            Bet.amtHistory.append(self.myBet)
            Bet.profitHistory.append(Bet.micro_game_profit)
            amount = Bet.amtHistory[-1]
            if amount == 'B':
                amount.lower()
            self.micro_game_profit.text = 'Micro-game Profit: ' + str(Bet.micro_game_profit)
            self.money_count.text = 'Money Count: ' + str(Bet.moneyCount)
            self.total_bets.text = '# of bets: ' + str(len(Bet.sideHistory))
            self.mySide = ''

        # Stats on the sidebar
        self.total_wins.text = 'P: ' + str(Board.player) + '     ' + 'B: ' + str(Board.banker)

        self.shoe_ti.text = 'Shoe TI: ' + str(Board.shoe_ti)
        self.pre_ti.text = 'Pre-TI: ' + str(Board.calc_pre_ti())

        # Print on ScrollView
        if last_winner == 'P':
            # To handle the format of the board
            # Length of these lines must be 17 for the purpose of the mistake() function
            if len(amount) > 2:
                self.board.text += '\n ' + amount + ' ' + last_winner + ' ' * 7 + checkmark + ' ' * 2
            elif len(amount) > 1:
                self.board.text += '\n ' + amount + ' ' * 3 + last_winner + ' ' * 7 + checkmark + ' '
            elif len(amount) > 0:
                self.board.text += '\n ' + amount + ' ' * 5 + last_winner + ' ' * 7 + checkmark
            else:
                self.board.text += '\n ' + amount + ' ' * 8 + last_winner + ' ' * 6  # 17
        elif last_winner == 'B':
            if len(amount) > 2:
                self.board.text += '\n ' + amount + ' ' * 7 + last_winner + ' ' + checkmark + ' ' * 2
            elif len(amount) > 1:
                self.board.text += '\n ' + amount + ' ' * 9 + last_winner + ' ' + checkmark + ' '
            elif len(amount) > 0:               # 1
                self.board.text += '\n ' + amount + ' ' * 11 + last_winner + ' ' + checkmark
            else:                               # 0
                self.board.text += '\n ' + amount + ' ' * 14 + last_winner  # 17

    def mistake(self):
        # The board cannot be empty
        if len(Board.history) != 0:

            # Deletes last winner from ScrollView
            self.board.text = self.board.text[:-17]

            # Undo last betting and profit make/lost from it
            if len(Bet.placeHistory) != 0 and Bet.placeHistory[-1] == len(Board.history) + 1:
                # last_bet = Bet.sideHistory[-1]
                # self.mySide = last_bet
                Bet.undo_profit()
                Bet.mistake()
                self.current_bet.text = self.myBet + ' bet on ?'
                self.micro_game_profit.text = 'Micro-game profit: ' + str(Bet.micro_game_profit)
                self.money_count.text = 'Money Count: ' + str(Bet.moneyCount)
                self.total_bets.text = '# of bets: ' + str(len(Bet.sideHistory))

            Board.mistake()
            # Stats on the sidebar
            self.total_wins.text = 'P: ' + str(Board.player) + '     ' + 'B: ' + str(Board.banker)
            self.shoe_ti.text = 'Shoe TI: ' + str(Board.shoe_ti)
            self.pre_ti.text = 'Pre-TI: ' + str(Board.calc_pre_ti())
        else:
            print('Board is empty!')

    @staticmethod
    def print_worksheet(worksheet):
        board = Board.history
        ti = Board.tiHistory
        bet_place = Bet.placeHistory
        bet_side = Bet.sideHistory
        bet_amt = Bet.amtHistory
        bet_profit = Bet.profitHistory
        micro_game_history = Bet.endMicroGameHistory

        worksheet["A1"] = "TI Count"
        worksheet['B1'] = "1A"
        worksheet["C1"] = "Player"
        worksheet["D1"] = "Banker"
        worksheet["E1"] = "Board"
        worksheet["F1"] = "Checks"
        worksheet["G1"] = "Profits"

        banker_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        player_fill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
        game_end = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        player_font = Font(color='FFFFFFFF')

        worksheet["C1"].fill = player_fill
        worksheet["D1"].fill = banker_fill

        for i in range(len(board)):
            # Shoe TI counts
            worksheet.cell(column=1, row=i + 2, value=ti[i])

            # Record betting side and amounts
            if i < len(bet_place):
                profit_cell = worksheet.cell(column=7, row=bet_place[i], value=bet_profit[i])
                if bet_side[i] == 'P':
                    bet_cell = worksheet.cell(column=3, row=bet_place[i], value=bet_amt[i])

                    # colors end of game
                    if (i + 1 < len(bet_place) and i in micro_game_history) or i + 1 == len(bet_place):
                        bet_cell.fill = game_end
                        profit_cell.fill = game_end
                elif bet_side[i] == 'B':
                    bet_cell = worksheet.cell(column=4, row=bet_place[i], value=bet_amt[i])

                    # colors end of game
                    if (i + 1 < len(bet_place) and i in micro_game_history) or i + 1 == len(bet_place):
                        bet_cell.fill = game_end
                        profit_cell.fill = game_end

                # Marks wins and losses
                checks = worksheet.cell(column=6, row=bet_place[i])
                checks.alignment = Alignment(horizontal='right')
                if bet_side[i] == board[bet_place[i] - 2]:
                    checks.value = '✓'
                elif bet_side[i] != board[bet_place[i] - 2]:
                    checks.value = 'x'

            # Marks results
            board_cell = worksheet.cell(column=5, row=i + 2, value=board[i])
            if board[i] == 'B':
                board_cell.alignment = Alignment(horizontal='right')
                board_cell.fill = banker_fill
            else:
                board_cell.fill = player_fill
                board_cell.font = player_font

        total_color = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

        # Marks total profits
        total_row = len(board) + 2
        total_cell = worksheet.cell(column=6, row=total_row, value='Total:')
        total_amt = worksheet.cell(column=7, row=total_row, value=Bet.moneyCount)
        total_cell.alignment = Alignment(horizontal='left')

        total_cell.fill = total_color
        total_amt.fill = total_color

    def export_to_excel(self):

        # When filename already exist
        try:
            wb = load_workbook(filename=filename)
            ws = wb.create_sheet("Sheet")
            self.print_worksheet(ws)
            wb.save(filename)
        # Create new file
        except FileNotFoundError:
            workbook = Workbook()
            worksheet = workbook.active
            self.print_worksheet(worksheet)
            workbook.save(filename)
        except PermissionError:
            print("Close the excel file first!")


class TICounterApp(App):
    title = "TICounter"

    def build(self):
        self.root = Screen()
        return self.root


if __name__ == "__main__":
    TICounterApp().run()